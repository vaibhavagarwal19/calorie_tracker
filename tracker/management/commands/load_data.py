import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from tracker.models import Food, Activity
from django.conf import settings


class Command(BaseCommand):
    help = 'Load food and activity data from Excel files'

    def handle(self, *args, **options):
        base_dir = settings.BASE_DIR
        self.load_foods(os.path.join(base_dir, 'data-excels-for-db', 'food-calories.xlsx'))
        self.load_activities(os.path.join(base_dir, 'data-excels-for-db', 'MET-values.xlsx'))

    def load_foods(self, filepath):
        self.stdout.write(f'Loading foods from {filepath}...')
        wb = load_workbook(filepath)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        foods = []
        for row in rows:
            if not row[0]:
                continue
            foods.append(Food(
                food_id=int(row[0]),
                name=str(row[1] or ''),
                food_group=str(row[2] or ''),
                calories=float(row[3] or 0),
                fat=float(row[4] or 0),
                protein=float(row[5] or 0),
                carbohydrate=float(row[6] or 0),
                serving_description=str(row[7] or ''),
            ))

        Food.objects.all().delete()
        Food.objects.bulk_create(foods, batch_size=1000)
        self.stdout.write(self.style.SUCCESS(f'Loaded {len(foods)} food items.'))

    def load_activities(self, filepath):
        self.stdout.write(f'Loading activities from {filepath}...')
        wb = load_workbook(filepath)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        activities = []
        for row in rows:
            if not row[0]:
                continue
            activities.append(Activity(
                activity_name=str(row[0] or ''),
                specific_motion=str(row[1] or ''),
                met_value=float(row[2] or 0),
            ))

        Activity.objects.all().delete()
        Activity.objects.bulk_create(activities, batch_size=500)
        self.stdout.write(self.style.SUCCESS(f'Loaded {len(activities)} activities.'))
